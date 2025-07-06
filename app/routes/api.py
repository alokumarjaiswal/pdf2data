from fastapi import APIRouter, HTTPException, Request, Response
from fastapi.responses import JSONResponse
import json
import base64
import os
import tempfile
import asyncio
from datetime import datetime
from pdf2image import convert_from_path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from app.db.mongo import parsed_collection, documents_collection, extractions_collection, processing_logs_collection, gridfs_bucket, LogManager, DocumentManager
from app.utils.data_lifecycle import DataLifecycleManager

router = APIRouter()

@router.get("/api/list")
async def get_parsed_documents_list():
    """Get list of all saved (approved) parsed documents with enhanced metadata"""
    cursor = parsed_collection.find(
        {"saved": True},  # Only show saved/approved documents
        {
            "_id": 1, 
            "parser": 1, 
            "original_filename": 1, 
            "uploaded_at": 1,
            "extraction_mode_used": 1,
            "num_entries": 1,
            "processing_completed": 1,
            "saved": 1,
            "saved_at": 1
        }
    )
    docs = await cursor.to_list(length=100)
    
    # Add processing stage information for each document
    for doc in docs:
        file_id = doc["_id"]
        doc_metadata = await documents_collection.find_one({"_id": file_id})
        if doc_metadata:
            doc["processing_stages"] = doc_metadata.get("processing_stages", {})
            doc["file_size"] = doc_metadata.get("file_size")
    
    # Convert datetime objects to strings for JSON serialization
    serializable_docs = json.loads(json.dumps(docs, default=str))
    return {"entries": serializable_docs}

@router.get("/api/data/{file_id}")
async def get_parsed_data(file_id: str, request: Request):
    """Get parsed data for a specific file with enhanced metadata"""
    # First try to get parsed data
    doc = await parsed_collection.find_one({"_id": file_id})
    
    if not doc:
        # If no parsed data, check if document exists in documents collection
        doc_metadata = await documents_collection.find_one({"_id": file_id})
        if not doc_metadata:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # Create a basic response for unparsed documents
        doc = {
            "_id": file_id,
            "file_id": file_id,
            "original_filename": doc_metadata.get("original_filename"),
            "uploaded_at": doc_metadata.get("uploaded_at"),
            "status": doc_metadata.get("status", "uploaded"),
            "processing_stages": doc_metadata.get("processing_stages", {}),
            "file_size": doc_metadata.get("file_size"),
            "parsed": False  # Indicate this document hasn't been parsed yet
        }
    else:
        # Keep _id for frontend use, but rename it to file_id for clarity
        doc["file_id"] = doc.pop("_id")
        doc["parsed"] = True  # Indicate this document has been parsed
    
    # Add additional metadata from other collections if not already present
    if "processing_stages" not in doc:
        doc_metadata = await documents_collection.find_one({"_id": file_id})
        if doc_metadata:
            doc["processing_stages"] = doc_metadata.get("processing_stages", {})
            doc["file_size"] = doc_metadata.get("file_size")
    
    # Add extraction information
    extraction_record = await extractions_collection.find_one({"file_id": file_id})
    if extraction_record:
        doc["extraction_info"] = {
            "mode": extraction_record.get("extraction_mode"),
            "method": extraction_record.get("method_used"),
            "pages": extraction_record.get("num_pages"),
            "chars": extraction_record.get("num_chars"),
            "extracted_at": extraction_record.get("extracted_at")
        }
    else:
        # If no extraction data yet, try to use page count from upload first
        doc_metadata = await documents_collection.find_one({"_id": file_id})
        if doc_metadata and doc_metadata.get("page_count"):
            # Use immediate page count from upload
            doc["extraction_info"] = {
                "mode": "upload_count",
                "method": "pdfplumber",
                "pages": doc_metadata["page_count"],
                "chars": 0,
                "extracted_at": None
            }
            print(f"📄 Using upload page count for {file_id}: {doc_metadata['page_count']} pages")
        elif doc_metadata and "gridfs_file_id" in doc_metadata:
            try:
                # Get PDF from GridFS and count pages
                pdf_file = await gridfs_bucket.open_download_stream(doc_metadata["gridfs_file_id"])
                pdf_content = await pdf_file.read()
                
                # Use pdf2image to count pages
                import tempfile
                from pdf2image import convert_from_path
                
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_file:
                    temp_file.write(pdf_content)
                    temp_pdf_path = temp_file.name
                
                try:
                    # Convert all pages to get count
                    pages = convert_from_path(temp_pdf_path, dpi=50)  # Low DPI just for counting
                    page_count = len(pages)
                    
                    doc["extraction_info"] = {
                        "mode": "direct_count",
                        "method": "pdf2image",
                        "pages": page_count,
                        "chars": 0,
                        "extracted_at": None
                    }
                    print(f"📄 Counted {page_count} pages for {file_id}")
                    
                finally:
                    import os
                    try:
                        os.unlink(temp_pdf_path)
                    except:
                        pass
                        
            except Exception as e:
                print(f"⚠️ Could not count pages for {file_id}: {e}")
                # Fallback to 1 page
                doc["extraction_info"] = {
                    "mode": "fallback",
                    "method": "assumed",
                    "pages": 1,
                    "chars": 0,
                    "extracted_at": None
                }

    pretty = request.query_params.get("pretty") == "1"

    if pretty:
        json_str = json.dumps(doc, indent=2, ensure_ascii=False, default=str)
        return Response(content=json_str, media_type="application/json; charset=utf-8")
    else:
        return JSONResponse(content=json.loads(json.dumps(doc, default=str)))

@router.post("/api/save/{file_id}")
async def save_parsed_document(file_id: str):
    """Save a parsed document to make it visible in the List Page"""
    try:
        # Check if the document exists in parsed_collection
        doc = await parsed_collection.find_one({"_id": file_id})
        if not doc:
            raise HTTPException(status_code=404, detail="Parsed document not found")
        
        # Check if already saved
        if doc.get("saved", False):
            return {
                "message": "Document already saved",
                "file_id": file_id,
                "saved": True,
                "saved_at": doc.get("saved_at")
            }
        
        # Update the document to mark it as saved
        save_timestamp = datetime.utcnow()
        
        await parsed_collection.update_one(
            {"_id": file_id},
            {
                "$set": {
                    "saved": True,
                    "saved_at": save_timestamp.isoformat()
                }
            }
        )
        
        # Update document processing stage to include saved
        await DocumentManager.update_processing_stage(file_id, "saved", True)
        
        # Log the save action
        await LogManager.store_log(
            file_id=file_id,
            process_type="save",
            log_content=f"Document saved and approved for listing: {doc.get('original_filename', 'Unknown.pdf')}",
            metadata={
                "parser": doc.get("parser"),
                "num_entries": doc.get("num_entries", 0),
                "saved_at": save_timestamp.isoformat()
            }
        )
        
        return {
            "message": "Document saved successfully",
            "file_id": file_id,
            "saved": True,
            "saved_at": save_timestamp.isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving document: {str(e)}")

@router.delete("/api/delete/{file_id}")
async def delete_parsed_document(file_id: str):
    """Delete a parsed document and all its associated data from database with enhanced validation"""
    try:
        # Use enhanced deletion with comprehensive validation and logging
        result = await DataLifecycleManager.enhanced_delete_document(file_id)
        
        if result["errors"]:
            raise HTTPException(status_code=500, detail=f"Deletion failed: {'; '.join(result['errors'])}")
        
        if not result["deleted_items"] and not result["warnings"]:
            raise HTTPException(status_code=404, detail="Document not found in any collection.")

        return {
            "message": "Document deleted successfully with enhanced validation",
            "file_id": file_id,
            "deleted_items": result["deleted_items"],
            "warnings": result["warnings"],
            "validation": result["validation"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting document: {str(e)}")

@router.get("/api/logs/{file_id}")
async def get_processing_logs(file_id: str, process_type: str = None):
    """Get processing logs for a specific file"""
    query = {"file_id": file_id}
    if process_type:
        query["process_type"] = process_type
    
    cursor = processing_logs_collection.find(query).sort("logged_at", 1)
    logs = await cursor.to_list(length=None)
    
    # Convert datetime objects to strings for JSON serialization
    serializable_logs = json.loads(json.dumps(logs, default=str))
    return {
        "file_id": file_id,
        "process_type": process_type,
        "logs": serializable_logs
    }

@router.get("/api/stats")
async def get_system_stats():
    """Get system statistics from database"""
    stats = {}
    
    # Document counts
    stats["total_documents"] = await documents_collection.count_documents({})
    stats["uploaded_documents"] = await documents_collection.count_documents({"processing_stages.uploaded": True})
    stats["extracted_documents"] = await documents_collection.count_documents({"processing_stages.extracted": True})
    stats["parsed_documents"] = await parsed_collection.count_documents({})
    
    # Extraction statistics
    extraction_stats = await extractions_collection.aggregate([
        {"$group": {
            "_id": "$extraction_mode",
            "count": {"$sum": 1},
            "avg_pages": {"$avg": "$num_pages"},
            "avg_chars": {"$avg": "$num_chars"}
        }}
    ]).to_list(length=None)
    stats["extraction_modes"] = extraction_stats
    
    # Parser statistics  
    parser_stats = await parsed_collection.aggregate([
        {"$group": {
            "_id": "$parser",
            "count": {"$sum": 1},
            "avg_entries": {"$avg": "$num_entries"}
        }}
    ]).to_list(length=None)
    stats["parsers_used"] = parser_stats
    
    # Convert datetime objects to strings for JSON serialization
    return json.loads(json.dumps(stats, default=str))

# New lifecycle management endpoints

@router.get("/api/lifecycle/orphaned")
async def get_orphaned_documents(max_age_hours: int = 168):
    """Get list of orphaned documents (uploaded/extracted but not parsed)"""
    try:
        orphaned_docs = await DataLifecycleManager.find_orphaned_documents(max_age_hours)
        return {
            "found": len(orphaned_docs),
            "max_age_hours": max_age_hours,
            "orphaned_documents": orphaned_docs
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error finding orphaned documents: {str(e)}")

@router.post("/api/lifecycle/cleanup")
async def cleanup_orphaned_documents(max_age_hours: int = 168, dry_run: bool = True):
    """Clean up orphaned documents older than specified hours"""
    try:
        if dry_run:
            orphaned_docs = await DataLifecycleManager.find_orphaned_documents(max_age_hours)
            return {
                "dry_run": True,
                "would_delete": len(orphaned_docs),
                "orphaned_documents": orphaned_docs,
                "message": "This was a dry run. Set dry_run=false to actually delete."
            }
        else:
            # Real cleanup - implement cleanup_orphaned_document method
            orphaned_docs = await DataLifecycleManager.find_orphaned_documents(max_age_hours)
            cleanup_results = []
            
            for doc in orphaned_docs:
                result = await DataLifecycleManager.enhanced_delete_document(doc["file_id"])
                cleanup_results.append(result)
            
            successful = sum(1 for r in cleanup_results if not r["errors"])
            failed = len(cleanup_results) - successful
            
            return {
                "dry_run": False,
                "total_found": len(orphaned_docs),
                "successfully_cleaned": successful,
                "failed_cleanups": failed,
                "cleanup_details": cleanup_results
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during cleanup: {str(e)}")

@router.get("/api/lifecycle/stats")
async def get_lifecycle_stats():
    """Get comprehensive data lifecycle statistics"""
    try:
        # Basic counts
        stats = {
            "total_documents": await documents_collection.count_documents({}),
            "total_extractions": await extractions_collection.count_documents({}),
            "total_parsed": await parsed_collection.count_documents({}),
            "total_logs": await processing_logs_collection.count_documents({})
        }
        
        # Processing stage analysis
        stats["workflow_stages"] = {
            "uploaded_only": await documents_collection.count_documents({
                "processing_stages.uploaded": True,
                "processing_stages.extracted": False,
                "processing_stages.parsed": False
            }),
            "extracted_not_parsed": await documents_collection.count_documents({
                "processing_stages.uploaded": True,
                "processing_stages.extracted": True,
                "processing_stages.parsed": False
            }),
            "fully_processed": await documents_collection.count_documents({
                "processing_stages.uploaded": True,
                "processing_stages.extracted": True,
                "processing_stages.parsed": True
            })
        }
        
        # Orphaned documents count (7 days instead of 24 hours)
        orphaned_docs = await DataLifecycleManager.find_orphaned_documents(168)  # Changed from 24 to 168 hours  
        stats["orphaned_documents_7d"] = len(orphaned_docs)  # Changed name to reflect 7 days
        
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error getting lifecycle stats: {str(e)}")

@router.post("/api/cleanup/failed-parsing")
async def cleanup_failed_parsing_results():
    """Clean up failed parsing results that were incorrectly saved to database"""
    try:
        # Find documents with failed parsing results
        failed_docs = []
        cursor = parsed_collection.find({})
        
        async for doc in cursor:
            if doc.get("tables"):
                for table in doc["tables"]:
                    if isinstance(table, dict) and table.get("success") is False:
                        failed_docs.append({
                            "file_id": doc["_id"],
                            "parser": doc.get("parser"),
                            "error": table.get("error", "Unknown error"),
                            "filename": doc.get("original_filename")
                        })
                        break
        
        # Remove failed parsing results
        cleanup_results = []
        for failed_doc in failed_docs:
            file_id = failed_doc["file_id"]
            
            # Delete from parsed_collection
            delete_result = await parsed_collection.delete_one({"_id": file_id})
            
            # Log the cleanup
            await LogManager.store_log(
                file_id=file_id,
                process_type="cleanup",
                log_content=f"Removed failed parsing result: {failed_doc['error']}",
                metadata={
                    "cleanup_type": "failed_parsing",
                    "parser": failed_doc["parser"],
                    "error": failed_doc["error"]
                }
            )
            
            cleanup_results.append({
                "file_id": file_id,
                "filename": failed_doc["filename"],
                "parser": failed_doc["parser"],
                "error": failed_doc["error"],
                "removed": delete_result.deleted_count > 0
            })
        
        return {
            "message": "Failed parsing results cleanup completed",
            "found_failed": len(failed_docs),
            "cleaned_up": len([r for r in cleanup_results if r["removed"]]),
            "results": cleanup_results
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error during cleanup: {str(e)}")

@router.get("/api/page-preview/{file_id}/{page_num}")
async def get_page_preview(file_id: str, page_num: int):
    """Get a preview image of a specific page from a PDF"""
    try:
        print(f"🔵 PAGE PREVIEW: Request for file_id={file_id}, page={page_num}")
        
        # Validate page number
        if page_num < 1:
            print(f"❌ PAGE PREVIEW: Invalid page number: {page_num}")
            raise HTTPException(status_code=400, detail="Page number must be >= 1")
            
        # Check if document exists in documents collection
        print(f"🔵 PAGE PREVIEW: Checking documents collection for {file_id}")
        doc = await documents_collection.find_one({"_id": file_id})
        if not doc:
            print(f"❌ PAGE PREVIEW: Document not found in documents collection: {file_id}")
            
            # Let's also check what documents DO exist
            all_docs = await documents_collection.find({}, {"_id": 1, "original_filename": 1}).to_list(length=10)
            print(f"📋 PAGE PREVIEW: Available documents in DB: {all_docs}")
            
            raise HTTPException(status_code=404, detail="Document not found")

        print(f"✅ PAGE PREVIEW: Document found: {doc.get('original_filename')}")
        print(f"🔵 PAGE PREVIEW: GridFS file_id: {doc.get('gridfs_file_id')}")

        # Check if file exists in GridFS
        try:
            print(f"🔵 PAGE PREVIEW: Opening GridFS stream for {file_id}")
            # Use the GridFS file ID from the document metadata, not the document file_id
            gridfs_file_id = doc.get("gridfs_file_id")
            if not gridfs_file_id:
                raise Exception("No GridFS file_id found in document metadata")
            
            print(f"🔵 PAGE PREVIEW: Using GridFS file_id: {gridfs_file_id}")
            pdf_file = await gridfs_bucket.open_download_stream(gridfs_file_id)
            pdf_content = await pdf_file.read()
            print(f"✅ PAGE PREVIEW: PDF content retrieved, size: {len(pdf_content)} bytes")
        except Exception as e:
            print(f"❌ PAGE PREVIEW: GridFS error for {file_id}: {e}")
            
            # Let's also check what files are in GridFS
            files_cursor = gridfs_bucket.find({"metadata.file_id": file_id})
            gridfs_files = await files_cursor.to_list(length=10)
            print(f"📋 PAGE PREVIEW: GridFS files for this file_id: {gridfs_files}")
            
            # Check all GridFS files
            all_files_cursor = gridfs_bucket.find({})
            all_gridfs_files = await all_files_cursor.to_list(length=10)
            print(f"📋 PAGE PREVIEW: All GridFS files: {[f.get('metadata', {}).get('file_id') for f in all_gridfs_files]}")
            
            raise HTTPException(status_code=404, detail="PDF file not found in storage")
        
        # Create temporary file
        temp_pdf_path = None
        try:
            print(f"🔵 PAGE PREVIEW: Creating temporary file")
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_file:
                temp_file.write(pdf_content)
                temp_pdf_path = temp_file.name
            
            print(f"✅ PAGE PREVIEW: Temporary file created: {temp_pdf_path}")

            # Convert the specific page to image (pdf2image uses 1-indexed pages)
            print(f"🔵 PAGE PREVIEW: Converting page {page_num} to image...")
            pages = convert_from_path(
                temp_pdf_path, 
                dpi=120,  # Reduced DPI for faster loading
                first_page=page_num, 
                last_page=page_num
            )
            
            if not pages:
                print(f"❌ PAGE PREVIEW: No pages found for page number {page_num}")
                raise HTTPException(status_code=404, detail=f"Page {page_num} not found")
            
            print(f"✅ PAGE PREVIEW: Page {page_num} converted successfully")
            
            # Convert to base64
            buffer = BytesIO()
            pages[0].save(buffer, format="PNG", optimize=True)
            image_b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            
            print(f"🎉 PAGE PREVIEW: Preview generated successfully for page {page_num}")
            
            return {
                "success": True,
                "page_num": page_num,
                "image": f"data:image/png;base64,{image_b64}",
                "file_id": file_id,
                "timestamp": asyncio.get_event_loop().time()  # Add timestamp to prevent caching
            }
            
        finally:
            # Clean up temporary file
            if temp_pdf_path and os.path.exists(temp_pdf_path):
                try:
                    os.unlink(temp_pdf_path)
                    print(f"✅ PAGE PREVIEW: Cleaned up temporary file: {temp_pdf_path}")
                except Exception as cleanup_error:
                    print(f"⚠️ PAGE PREVIEW: Failed to cleanup temp file {temp_pdf_path}: {cleanup_error}")
                    
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        print(f"❌ PAGE PREVIEW: Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating page preview: {str(e)}")

@router.put("/api/update/{file_id}")
async def update_parsed_document(file_id: str, request: Request):
    """Update the parsed data for a specific document"""
    try:
        # Get the updated data from request body
        updated_data = await request.json()
        
        # Check if the document exists in parsed_collection
        existing_doc = await parsed_collection.find_one({"_id": file_id})
        if not existing_doc:
            raise HTTPException(status_code=404, detail="Parsed document not found")
        
        # Preserve essential metadata while updating the data
        essential_fields = {
            "_id": file_id,
            "parser": existing_doc.get("parser"),
            "original_filename": existing_doc.get("original_filename"),
            "uploaded_at": existing_doc.get("uploaded_at"),
            "extraction_mode_used": existing_doc.get("extraction_mode_used"),
            "processing_completed": existing_doc.get("processing_completed"),
            "saved": existing_doc.get("saved", False),
            "saved_at": existing_doc.get("saved_at")
        }
        
        # Merge updated data with essential fields
        merged_data = {**updated_data, **essential_fields}
        
        # Update the document in the collection
        update_timestamp = datetime.utcnow()
        merged_data["last_modified"] = update_timestamp.isoformat()
        
        await parsed_collection.replace_one(
            {"_id": file_id},
            merged_data
        )
        
        # Log the update action
        await LogManager.store_log(
            file_id=file_id,
            process_type="update",
            log_content=f"Document data updated: {existing_doc.get('original_filename', 'Unknown.pdf')}",
            metadata={
                "parser": existing_doc.get("parser"),
                "modified_at": update_timestamp.isoformat(),
                "updated_by": "user_edit"
            }
        )
        
        return {
            "message": "Document updated successfully",
            "file_id": file_id,
            "last_modified": update_timestamp.isoformat()
        }
        
    except HTTPException:
        raise
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in request body")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating document: {str(e)}")

@router.get("/api/file/{file_id}")
async def get_file(file_id: str):
    """Serve PDF file for viewing"""
    from fastapi.responses import FileResponse
    import os
    
    try:
        # Try to serve from local directory first
        file_path = f"data/uploaded_pdfs/original_{file_id}.pdf"
        if os.path.exists(file_path):
            return FileResponse(
                path=file_path,
                media_type="application/pdf",
                headers={"Content-Disposition": "inline"}
            )
        
        # If not found locally, try to get from GridFS
        doc = await documents_collection.find_one({"_id": file_id})
        if not doc:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Get file from GridFS
        try:
            file_data = await gridfs_bucket.download_to_stream(file_id)
            # Create a temporary file to serve
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_file:
                temp_file.write(file_data.getvalue())
                temp_path = temp_file.name
            
            return FileResponse(
                path=temp_path,
                media_type="application/pdf",
                headers={"Content-Disposition": "inline"},
                background=None  # Don't auto-delete the temp file
            )
        except Exception as gridfs_error:
            raise HTTPException(status_code=404, detail="File not found in storage")
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error serving file: {str(e)}")

@router.get("/api/extracted-text/{file_id}")
async def get_extracted_text(file_id: str):
    """Serve extracted text file for viewing"""
    from fastapi.responses import PlainTextResponse
    import os
    
    try:
        # Look for the extracted text file
        file_path = f"data/extracted_pages/extracted_digital_{file_id}.txt"
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return PlainTextResponse(
                content=content,
                headers={"Content-Disposition": "inline"}
            )
        else:
            raise HTTPException(status_code=404, detail="Extracted text not found")
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error serving extracted text: {str(e)}")

@router.post("/api/unite/upload/{file_id}")
async def upload_to_unite(file_id: str):
    """Queue document upload to Unite ERP system"""
    import asyncio
    import subprocess
    import os
    
    try:
        # Check if document exists
        doc = await parsed_collection.find_one({"_id": file_id})
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        # TODO: In a production system, you'd want to use a proper task queue like Celery
        # For now, we'll return success and let the frontend handle the UI
        
        # Update document status to indicate Unite upload is pending
        await parsed_collection.update_one(
            {"_id": file_id},
            {"$set": {
                "unite_status": "uploading",
                "unite_upload_initiated_at": datetime.utcnow()
            }}
        )
        
        # In the future, you could trigger the bot script here:
        # bot_path = os.path.join(os.path.dirname(__file__), "../../unite-login-bot/login.py")
        # subprocess.Popen(["python", bot_path, file_id], cwd=os.path.dirname(bot_path))
        
        return {
            "success": True,
            "message": "Upload queued successfully",
            "file_id": file_id,
            "status": "uploading"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        # Update status to error
        await parsed_collection.update_one(
            {"_id": file_id},
            {"$set": {
                "unite_status": "error",
                "unite_error": str(e),
                "unite_error_at": datetime.utcnow()
            }}
        )
        raise HTTPException(status_code=500, detail=f"Failed to queue upload: {str(e)}")

@router.get("/api/unite/status/{file_id}")
async def get_unite_status(file_id: str):
    """Get the Unite upload status for a document"""
    try:
        doc = await parsed_collection.find_one(
            {"_id": file_id},
            {"unite_status": 1, "unite_uploaded_at": 1, "unite_error": 1}
        )
        
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        return {
            "file_id": file_id,
            "unite_status": doc.get("unite_status", "pending"),
            "unite_uploaded_at": doc.get("unite_uploaded_at"),
            "unite_error": doc.get("unite_error")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get status: {str(e)}")

@router.get("/api/export/excel/{file_id}")
async def export_to_excel(file_id: str):
    """Export parsed data to Excel format"""
    try:
        # Validate file_id format
        if not file_id or len(file_id.strip()) == 0:
            raise HTTPException(status_code=400, detail="Invalid file ID provided")
        
        # Get parsed data from database
        doc = await parsed_collection.find_one({"_id": file_id})
        
        if not doc:
            raise HTTPException(
                status_code=404, 
                detail=f"Parsed document with ID '{file_id}' not found. Please ensure the document has been processed and saved."
            )
        
        if not doc.get("saved", False):
            raise HTTPException(
                status_code=400, 
                detail="Document must be saved before exporting. Please save the document first and try again."
            )
        
        # Get document metadata for filename
        original_filename = doc.get("original_filename", "Unknown.pdf")
        parser_used = doc.get("parser", "Unknown")
        tables = doc.get("tables", [])
        
        if not tables:
            raise HTTPException(
                status_code=400, 
                detail="No parsed data available for export. Please ensure the document has been parsed successfully."
            )
        
        # Validate that tables contain actual data
        has_data = False
        for table in tables:
            if isinstance(table, dict) and table:
                has_data = True
                break
        
        if not has_data:
            raise HTTPException(
                status_code=400,
                detail="No valid data found in parsed tables. Please re-parse the document."
            )
        
        # Create Excel workbook
        try:
            wb = Workbook()
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to create Excel workbook: {str(e)}"
            )
        
        # Define styles
        try:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to create Excel styles: {str(e)}"
            )
        
        if parser_used == "DaybookParser":
            # Handle Daybook data structure
            ws = wb.active
            ws.title = "Daybook Data"
            
            # Track current row
            current_row = 1
            
            for table_idx, table in enumerate(tables):
                # Society header
                ws.merge_cells(f'A{current_row}:E{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = f"Society: {table.get('society_name', 'N/A')}"
                cell.font = Font(bold=True, size=14)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = center_alignment
                current_row += 1
                
                # Village and Date info
                ws[f'A{current_row}'] = "Village:"
                ws[f'B{current_row}'] = table.get('village_name', 'N/A')
                ws[f'C{current_row}'] = "Date:"
                ws[f'D{current_row}'] = table.get('date', 'N/A')
                current_row += 2
                
                # Entries table headers
                headers = ['Account Name', 'Account Number', 'Amount', 'Total Amount', 'Type']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                current_row += 1
                
                # Entries data
                entries = table.get('entries', [])
                for entry in entries:
                    ws.cell(row=current_row, column=1, value=entry.get('account_name', '')).border = border
                    ws.cell(row=current_row, column=2, value=entry.get('account_number', '')).border = border
                    
                    amount = entry.get('amount')
                    if amount is not None:
                        ws.cell(row=current_row, column=3, value=float(amount)).border = border
                        ws.cell(row=current_row, column=3, value=float(amount)).number_format = '#,##0.00'
                    else:
                        ws.cell(row=current_row, column=3, value="").border = border
                    
                    total_amount = entry.get('total_amount')
                    if total_amount is not None:
                        ws.cell(row=current_row, column=4, value=float(total_amount)).border = border
                        ws.cell(row=current_row, column=4, value=float(total_amount)).number_format = '#,##0.00'
                    else:
                        ws.cell(row=current_row, column=4, value="").border = border
                    
                    # Determine entry type based on available data
                    entry_type = "Entry"
                    if entry.get('account_number'):
                        entry_type = "Account"
                    elif 'S/O' in entry.get('account_name', ''):
                        entry_type = "Person"
                    elif amount is not None and amount > 0:
                        entry_type = "Transaction"
                    
                    ws.cell(row=current_row, column=5, value=entry_type).border = border
                    current_row += 1
                
                # Totals section
                current_row += 1
                totals = table.get('totals', {})
                if totals:
                    # Totals header
                    ws.merge_cells(f'A{current_row}:E{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = "TOTALS"
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.alignment = center_alignment
                    current_row += 1
                    
                    # Total items
                    for key, value in totals.items():
                        if value is not None:
                            ws.cell(row=current_row, column=1, value=key.replace('_', ' ').title())
                            ws.cell(row=current_row, column=2, value=float(value))
                            ws.cell(row=current_row, column=2).number_format = '#,##0.00'
                            current_row += 1
                
                # Amount in words
                amount_in_words = table.get('amount_in_words', '')
                if amount_in_words:
                    current_row += 1
                    ws.merge_cells(f'A{current_row}:E{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = f"Amount in Words: {amount_in_words}"
                    cell.font = Font(italic=True)
                    cell.alignment = Alignment(horizontal='left')
                
                current_row += 3  # Add space between tables
            
            # Auto-adjust column widths
            column_widths = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.coordinate and cell.value:
                        col_letter = cell.column_letter
                        if col_letter not in column_widths:
                            column_widths[col_letter] = 0
                        try:
                            width = len(str(cell.value))
                            if width > column_widths[col_letter]:
                                column_widths[col_letter] = width
                        except:
                            pass
            
            for col_letter, width in column_widths.items():
                adjusted_width = min(width + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        else:
            # Handle other parser types generically
            ws = wb.active
            ws.title = f"{parser_used} Data"
            
            # Add summary sheet first
            current_row = 1
            
            # Document info header
            ws.merge_cells(f'A{current_row}:D{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f"Document: {original_filename} (Parser: {parser_used})"
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=14, color="FFFFFF")
            cell.alignment = center_alignment
            current_row += 2
            
            # Process each table
            for table_idx, table in enumerate(tables):
                ws.merge_cells(f'A{current_row}:D{current_row}')
                cell = ws[f'A{current_row}']
                cell.value = f"Table {table_idx + 1}"
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = center_alignment
                current_row += 1
                
                # Convert table data to key-value pairs
                if isinstance(table, dict):
                    for key, value in table.items():
                        ws.cell(row=current_row, column=1, value=str(key).replace('_', ' ').title())
                        
                        # Handle different value types
                        if isinstance(value, (list, dict)):
                            ws.cell(row=current_row, column=2, value=str(value)[:100] + "..." if len(str(value)) > 100 else str(value))
                        elif isinstance(value, (int, float)):
                            ws.cell(row=current_row, column=2, value=value)
                        else:
                            ws.cell(row=current_row, column=2, value=str(value) if value is not None else "")
                        
                        current_row += 1
                
                current_row += 2  # Space between tables
            
            # Auto-adjust column widths
            column_widths = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.coordinate and cell.value:
                        col_letter = cell.column_letter
                        if col_letter not in column_widths:
                            column_widths[col_letter] = 0
                        try:
                            width = len(str(cell.value))
                            if width > column_widths[col_letter]:
                                column_widths[col_letter] = width
                        except:
                            pass
            
            for col_letter, width in column_widths.items():
                adjusted_width = min(width + 2, 80)
                ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO object
        try:
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Get the content and verify that data was written
            content = excel_buffer.getvalue()
            if len(content) == 0:
                raise Exception("No data was written to Excel file")
                
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to generate Excel file: {str(e)}"
            )
        
        # Generate filename safely
        try:
            safe_filename = original_filename.replace('.pdf', '').replace(' ', '_').replace('/', '_').replace('\\', '_')
            # Remove any potentially problematic characters
            import re
            safe_filename = re.sub(r'[<>:"|?*]', '_', safe_filename)
            excel_filename = f"{safe_filename}_{parser_used}_data.xlsx"
        except Exception as e:
            # Fallback filename
            excel_filename = f"document_{parser_used}_data.xlsx"
        
        # Return Excel file with comprehensive headers
        try:
            # Final validation
            if len(content) == 0:
                raise HTTPException(
                    status_code=500,
                    detail="Generated Excel file is empty"
                )
            
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=\"{excel_filename}\"",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Length": str(len(content)),
                    "Cache-Control": "no-cache, no-store, must-revalidate",
                    "Pragma": "no-cache",
                    "Expires": "0"
                }
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to prepare Excel file for download: {str(e)}"
            )
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is (they already have proper status codes and messages)
        raise
    except Exception as e:
        # Log the full error for debugging
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌ Excel Export Error: {error_trace}")
        
        # Categorize errors for better user experience
        error_message = str(e)
        if "memory" in error_message.lower() or "memoryerror" in str(type(e)).lower():
            raise HTTPException(
                status_code=500, 
                detail="Document is too large to export. Please contact support."
            )
        elif "permission" in error_message.lower() or "access" in error_message.lower():
            raise HTTPException(
                status_code=500,
                detail="Server permission error. Please try again later."
            )
        elif "disk" in error_message.lower() or "space" in error_message.lower():
            raise HTTPException(
                status_code=500,
                detail="Server storage issue. Please try again later."
            )
        else:
            # Generic error with sanitized message
            raise HTTPException(
                status_code=500, 
                detail=f"Failed to export Excel file. Please try again or contact support if the problem persists."
            )
